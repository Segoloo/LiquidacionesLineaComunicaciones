#!/usr/bin/env python3
"""
Script para extraer información de liquidaciones de técnicos desde MySQL
Lee datos de data_linea_todos + facturacion_linea y genera JSON con resumen
MODIFICADO: 
- ENERO 2026: Lee desde Excel data/ENERO2026.xlsx
- FEBRERO 2026+: Extrae de base de datos
- Aplica descuentos por tipología
- Genera JSON y Excel resumen

CORRECCIÓN CRÍTICA:
- El Excel de entrada NO tiene descuentos aplicados (valores brutos)
- Los descuentos se aplican según la tipología AQUÍ en el script
- Los valores netos se calculan como: valor_bruto * (1 - descuento_porcentaje)
"""

import mysql.connector
from mysql.connector import Error
import json
import gzip
from datetime import datetime
from typing import Dict, List, Optional
import sys
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Configuración de la base de datos
DB_CONFIG = {
    'host': '192.168.4.92',
    'port': 3306,
    'user': 'root',
    'password': 'An4l1t1c4l1n34*',
    'database': 'lineacom_analitica'
}

# Meta mensual de recaudo (debe ser igual a app.js)
META_MENSUAL = 4500000  # $4,500,000

# Tramos de comisión (debe ser igual a app.js)
COMMISSION_TIERS = [
    {'min': 0, 'max': 1000000, 'percentage': 0.15, 'name': 'Uno'},
    {'min': 1000001, 'max': 2000000, 'percentage': 0.08, 'name': 'Dos'},
    {'min': 2000001, 'max': 3000000, 'percentage': 0.05, 'name': 'Tres'},
    {'min': 3000001, 'max': float('inf'), 'percentage': 0.03, 'name': 'Cuatro'}
]

# Configuración de descuentos por tipología (según imagen adjunta)
DESCUENTOS = {
    'TIPO I': 0.20,      # 20%
    'TIPO II': 0.30,     # 30%
    'TIPO III': 0.50,    # 50%
    'TIPO IV': 0.50,     # 50%
    'TIPO V': 0.60,      # 60%
    'PRINCIPAL': 0.20,   # 20%
    'INTERMEDIA': 0.30,  # 30%
    'LEJANA': 0.50       # 50%
}

class ExtractorLiquidacionesDB:
    """Clase para extraer y procesar liquidaciones desde la base de datos"""
    
    def __init__(self, db_config: dict):
        self.db_config = db_config
        self.connection = None
        self.liquidaciones = {}
    
    def calcular_comision(self, total_recaudado: float) -> dict:
        """
        Calcula la comisión por TRAMOS PROGRESIVOS (igual que app.js)
        
        Se calcula sumando la comisión de cada tramo que se alcance.
        Ejemplo: Si excedente es $1,500,000:
        - Primer millón × 15% = $150,000
        - $500,000 restantes × 8% = $40,000
        - TOTAL = $190,000
        
        Args:
            total_recaudado: Total recaudado por el técnico en el mes
            
        Returns:
            Diccionario con comisión, tier, excedente y si cumplió meta
        """
        # Si no alcanza la meta mínima, comisión es 0
        if total_recaudado < META_MENSUAL:
            return {
                'commission': 0,
                'tier': {'name': 'Ninguno', 'percentage': 0},
                'excedente': 0,
                'meta_cumplida': False
            }
        
        excedente = total_recaudado - META_MENSUAL
        commission = 0
        current_tier = None
        
        # Calcular comisión por tramos progresivos (igual que app.js)
        for tier in COMMISSION_TIERS:
            if excedente >= tier['min']:
                monto_en_tier = min(excedente - tier['min'], tier['max'] - tier['min'])
                if monto_en_tier > 0:
                    commission += monto_en_tier * tier['percentage']
                    current_tier = tier
        
        return {
            'commission': round(commission),
            'tier': current_tier if current_tier else COMMISSION_TIERS[0],
            'excedente': excedente,
            'meta_cumplida': True
        }
        
    def conectar(self) -> bool:
        """
        Establece conexión con la base de datos MySQL
        
        Returns:
            True si la conexión fue exitosa, False en caso contrario
        """
        try:
            self.connection = mysql.connector.connect(**self.db_config)
            if self.connection.is_connected():
                db_info = self.connection.server_info  # Usar propiedad en vez de método
                print(f"✓ Conectado a MySQL Server versión {db_info}")
                cursor = self.connection.cursor()
                cursor.execute("SELECT DATABASE();")
                record = cursor.fetchone()
                print(f"✓ Conectado a la base de datos: {record[0]}")
                cursor.close()
                return True
        except Error as e:
            print(f"✗ Error al conectar a MySQL: {e}")
            return False
    
    def desconectar(self):
        """Cierra la conexión con la base de datos"""
        if self.connection and self.connection.is_connected():
            self.connection.close()
            print("✓ Conexión a MySQL cerrada")
    
    def obtener_tablas_disponibles(self) -> List[str]:
        """
        Obtiene la lista de tablas disponibles en la base de datos
        
        Returns:
            Lista con nombres de tablas
        """
        try:
            cursor = self.connection.cursor()
            cursor.execute("SHOW TABLES")
            tablas = [tabla[0] for tabla in cursor.fetchall()]
            cursor.close()
            return tablas
        except Error as e:
            print(f"✗ Error al obtener tablas: {e}")
            return []
    
    def extraer_tareas_enero_desde_excel(self, ruta_excel: str = 'data/ENERO2026.xlsx') -> List[Dict]:
        """
        Extrae tareas de ENERO 2026 desde el archivo Excel
        
        CORRECCIÓN IMPORTANTE:
        - El Excel tiene valores BRUTOS (sin descuentos aplicados)
        - Los descuentos se calculan según la tipología
        - Se retorna valor_bruto, descuento_aplicado y prod_tecnico_final (neto)
        
        Args:
            ruta_excel: Ruta al archivo Excel de enero
            
        Returns:
            Lista de diccionarios con información de tareas
        """
        try:
            print(f"\n→ Leyendo archivo Excel: {ruta_excel}")
            
            # Leer hoja DATOS_COMPLETOS que contiene todas las tareas
            df = pd.read_excel(ruta_excel, sheet_name='DATOS_COMPLETOS')
            print(f"✓ Leídas {len(df)} filas del Excel")
            
            # Mostrar columnas encontradas para debug
            print(f"\n→ Columnas encontradas en el Excel:")
            columnas_tecnico = [c for c in df.columns if 'TECNICO' in str(c).upper()]
            columnas_tipif = [c for c in df.columns if 'TIPIF' in str(c).upper()]
            columnas_actividad = [c for c in df.columns if 'ACTIVIDAD' in str(c).upper()]
            
            if columnas_tecnico:
                print(f"  - Técnico: {', '.join(columnas_tecnico)}")
            else:
                print("  ⚠ NO se encontró columna de TECNICO")
            
            if columnas_tipif:
                print(f"  - Tipología: {', '.join(columnas_tipif)}")
            else:
                print("  ⚠ NO se encontró columna de TIPIFICACIÓN")
            
            if columnas_actividad:
                print(f"  - Actividad: {', '.join(columnas_actividad)}")
            
            # Mapeo de columnas del Excel a la estructura esperada
            columnas_mapeo = {
                'NOMBRE DEL TECNICO': 'tecnico',
                'NOMBRE  DEL TECNICO': 'tecnico',  # puede tener doble espacio
                'TIPIFICACION': 'tipologia',
                'ACTIVIDAD': 'tipo_actividad',
                'TA/CODIGO ACTIVIDAD': 'tarea',
                'CODIGO CB': 'codigo_sitio',
                'DEPARTAMENTO': 'departamento',
                'NOMBRE DEL CB': 'nombre_punto',
                'NOMBRE  DEL CB': 'nombre_punto',
                'CENTRO DE COSTOS LINEACOM': 'formulario',
                'ESTADO': 'estado_actividad',
                'FECHA CIERRE DEL TICKET (DD/MM/AAAA HH:MM)': 'fecha_cierre',
                'ZONA': 'zona_coordinador',
                'CIUDAD SEDE': 'ciudad_sede',
                'FORMATO': 'formato',
                'CADENA': 'cadena',
                'FORMA DE ATENCION': 'forma_atencion',
                'TIPO DE TRAYECTO': 'trayecto'
            }
            
            # Buscar columna de precio/valor - SIEMPRE es valor BRUTO
            col_precio = None
            
            # Buscar columnas de valor/precio (valores BRUTOS)
            for col in df.columns:
                col_upper = str(col).upper()
                if any(precio in col_upper for precio in ['VALOR TOTAL', 'VALOR DIAS ACTIVIDAD', 'PRECIO', 'VALOR']):
                    # Evitar columnas con descuentos
                    if not any(keyword in col_upper for keyword in ['DESCUENTO', 'NETO', 'TECNICO-DESCUENTO', 'ACTIVIDAD-DESCUENTO']):
                        col_precio = col
                        print(f"✓ Columna de precio BRUTO encontrada: {col_precio}")
                        break
            
            if not col_precio:
                print("✗ No se encontró columna de precio en el Excel")
                return []
            
            # IMPORTANTE: Los valores son BRUTOS (sin descuentos)
            print(f"\n{'='*70}")
            print("✓  Los valores en el Excel son BRUTOS (sin descuentos)")
            print("   → SE aplicarán descuentos por tipología según configuración")
            print(f"{'='*70}\n")
            
            # Convertir DataFrame a lista de diccionarios
            resultados = []
            tareas_procesadas = 0
            tareas_sin_tecnico = 0
            tareas_sin_valor = 0
            
            for idx, row in df.iterrows():
                # Encontrar nombre del técnico - buscar en múltiples columnas posibles
                tecnico = None
                for col in df.columns:
                    col_upper = str(col).upper()
                    # Buscar cualquier columna que contenga "TECNICO"
                    if 'TECNICO' in col_upper and pd.notna(row[col]):
                        tecnico_value = str(row[col]).strip()
                        if tecnico_value and tecnico_value.upper() not in ['', 'NONE', 'NAN']:
                            tecnico = tecnico_value
                            break
                
                if not tecnico or tecnico == '':
                    tareas_sin_tecnico += 1
                    continue
                
                # Encontrar tipología - buscar en múltiples columnas posibles
                tipologia = None
                for col in df.columns:
                    col_upper = str(col).upper()
                    if 'TIPIF' in col_upper and pd.notna(row[col]):
                        tipologia = str(row[col]).strip().upper()
                        break
                
                # Obtener valor BRUTO del Excel
                valor_bruto = 0
                if pd.notna(row[col_precio]):
                    try:
                        valor_bruto = float(row[col_precio])
                    except:
                        valor_bruto = 0
                
                if valor_bruto <= 0:
                    tareas_sin_valor += 1
                    continue
                
                # APLICAR DESCUENTO según tipología
                descuento_porcentaje = DESCUENTOS.get(tipologia, 0)
                valor_descuento = valor_bruto * descuento_porcentaje
                valor_neto = valor_bruto - valor_descuento
                
                # Obtener fecha de cierre - buscar columna de fecha
                fecha_cierre = None
                for col in df.columns:
                    col_upper = str(col).upper()
                    if 'FECHA' in col_upper and 'CIERRE' in col_upper:
                        fecha_str = row[col]
                        if pd.notna(fecha_str):
                            try:
                                # Intentar parsear la fecha
                                fecha_cierre = pd.to_datetime(fecha_str, format='%d/%m/%Y %H:%M', errors='coerce')
                                if pd.isna(fecha_cierre):
                                    # Intentar otros formatos
                                    fecha_cierre = pd.to_datetime(fecha_str, errors='coerce')
                                    if pd.isna(fecha_cierre):
                                        fecha_cierre = datetime(2026, 1, 15)
                            except:
                                fecha_cierre = datetime(2026, 1, 15)
                        break
                
                if not fecha_cierre:
                    fecha_cierre = datetime(2026, 1, 15)
                
                # Buscar tipo de actividad
                tipo_actividad = None
                for col in df.columns:
                    col_upper = str(col).upper()
                    if 'ACTIVIDAD' in col_upper and 'CODIGO' not in col_upper and pd.notna(row[col]):
                        tipo_actividad = str(row[col]).strip()
                        break
                
                # Buscar código de tarea
                tarea_codigo = None
                for col in df.columns:
                    col_upper = str(col).upper()
                    if ('TAREA' in col_upper or 'CODIGO' in col_upper and 'ACTIVIDAD' in col_upper) and pd.notna(row[col]):
                        tarea_codigo = str(row[col]).strip()
                        break
                
                # Crear diccionario con estructura similar a la DB
                tarea_dict = {
                    'tecnico': tecnico,
                    'tipologia': tipologia,
                    'valor_bruto': valor_bruto,  # Valor original del Excel (BRUTO)
                    'descuento_aplicado': valor_descuento,  # Descuento calculado
                    'prod_tecnico_final': valor_neto,  # Valor después de descuento (NETO)
                    'total_facturacion': valor_bruto,
                    'valor_total_entidad': valor_bruto,
                    'valor_total_red': valor_bruto,
                    'fecha_cierre_plataforma_cliente': fecha_cierre,  # Como objeto datetime
                    'tipo_actividad': tipo_actividad or '',
                    'tarea': tarea_codigo or '',
                }
                
                # Mapear otras columnas de forma flexible
                for col_excel, col_destino in columnas_mapeo.items():
                    if col_excel in df.columns and pd.notna(row[col_excel]):
                        tarea_dict[col_destino] = str(row[col_excel]).strip()
                
                # Campos adicionales que pueden no estar en el mapeo
                if 'bodega' not in tarea_dict:
                    # Buscar columna de ZONA
                    for col in df.columns:
                        if 'ZONA' in str(col).upper() and pd.notna(row[col]):
                            tarea_dict['bodega'] = str(row[col]).strip()
                            tarea_dict['zona_coordinador'] = str(row[col]).strip()
                            break
                    if 'bodega' not in tarea_dict:
                        tarea_dict['bodega'] = 'SIN ZONA'
                        tarea_dict['zona_coordinador'] = 'SIN ZONA'
                
                # Buscar ciudad
                if 'ciudad' not in tarea_dict:
                    for col in df.columns:
                        if 'CIUDAD' in str(col).upper() and 'SEDE' not in str(col).upper() and pd.notna(row[col]):
                            tarea_dict['ciudad'] = str(row[col]).strip()
                            break
                
                # Valores por defecto para campos faltantes
                tarea_dict.setdefault('departamento_completo', tarea_dict.get('departamento', ''))
                tarea_dict.setdefault('region_sitio', None)
                tarea_dict.setdefault('estado_ta', tarea_dict.get('estado_actividad', ''))
                tarea_dict.setdefault('estado_fo', None)
                tarea_dict.setdefault('resultado_actividad', None)
                
                resultados.append(tarea_dict)
                tareas_procesadas += 1
            
            print(f"\n→ Resumen de procesamiento:")
            print(f"  ✓ Tareas procesadas: {tareas_procesadas}")
            if tareas_sin_tecnico > 0:
                print(f"  ⚠ Tareas sin técnico: {tareas_sin_tecnico}")
            if tareas_sin_valor > 0:
                print(f"  ⚠ Tareas sin valor: {tareas_sin_valor}")
            
            # Mostrar ejemplo de descuentos aplicados
            if resultados:
                print(f"\n→ Ejemplo de tarea con descuentos aplicados:")
                ejemplo = resultados[0]
                print(f"  Técnico: {ejemplo['tecnico']}")
                print(f"  Tipología: {ejemplo.get('tipologia', 'N/A')}")
                print(f"  Valor bruto: ${ejemplo['valor_bruto']:,.2f}")
                print(f"  Descuento: ${ejemplo['descuento_aplicado']:,.2f} ({DESCUENTOS.get(ejemplo.get('tipologia', ''), 0)*100}%)")
                print(f"  Valor neto: ${ejemplo['prod_tecnico_final']:,.2f}")
            
            return resultados
            
        except FileNotFoundError:
            print(f"✗ No se encontró el archivo: {ruta_excel}")
            print(f"  Asegúrate de que el archivo existe en la ruta especificada")
            return []
        except Exception as e:
            print(f"✗ Error al leer Excel: {e}")
            import traceback
            traceback.print_exc()
            return []
    
    def extraer_tareas(self) -> List[Dict]:
        """
        Extrae todas las tareas de FEBRERO-DICIEMBRE 2026 desde las tablas de MySQL
        
        Returns:
            Lista de diccionarios con información de tareas
        """
        try:
            cursor = self.connection.cursor(dictionary=True)
            
            # Query OPTIMIZADA: une data_linea_todos y facturacion_linea
            # Filtra solo tareas del año 2026, meses 2-12 (febrero a diciembre)
            query = """
            SELECT 
                d.tecnico,
                d.tarea,
                d.ciudad,
                d.departamento_completo,
                d.bodega,
                d.nombre_punto,
                d.estado_ta,
                d.estado_fo,
                d.resultado_actividad,
                d.tipo_actividad,
                d.formulario,
                d.tipologia,
                d.region_sitio,
                d.fecha_cierre_plataforma_cliente,
                d.fecha_fin,
                f.prod_tecnico_final,
                f.total_facturacion,
                f.valor_total_entidad,
                f.valor_total_red,
                d.trayecto,
                d.codigo_sitio
            FROM data_linea_todos d
            LEFT JOIN facturacion_linea f ON d.tarea = f.tarea
            WHERE YEAR(d.fecha_cierre_plataforma_cliente) = 2026
              AND MONTH(d.fecha_cierre_plataforma_cliente) BETWEEN 2 AND 12
              AND d.tecnico IS NOT NULL
              AND d.tecnico != ''
            ORDER BY d.fecha_cierre_plataforma_cliente DESC
            """
            
            print(f"\n→ Ejecutando query para extraer tareas de Feb-Dic 2026...")
            cursor.execute(query)
            
            # Obtener todas las filas
            tareas = cursor.fetchall()
            cursor.close()
            
            if not tareas:
                print("✗ No se encontraron tareas en el rango especificado")
                return []
            
            print(f"✓ Extraídas {len(tareas)} tareas de la base de datos (Feb-Dic 2026)")
            
            # Procesar tareas para calcular descuentos
            tareas_procesadas = []
            for tarea in tareas:
                tipologia = str(tarea.get('tipologia', '') or '').strip().upper()
                
                # Obtener valor bruto (prod_tecnico_final es el valor sin descuento en la BD)
                valor_bruto = float(tarea.get('prod_tecnico_final', 0) or 0)
                
                if valor_bruto <= 0:
                    continue
                
                # Aplicar descuento según tipología
                descuento_porcentaje = DESCUENTOS.get(tipologia, 0)
                valor_descuento = valor_bruto * descuento_porcentaje
                valor_neto = valor_bruto - valor_descuento
                
                # Agregar campos calculados
                tarea['valor_bruto'] = valor_bruto
                tarea['descuento_aplicado'] = valor_descuento
                tarea['prod_tecnico_final'] = valor_neto  # Sobrescribir con valor neto
                
                tareas_procesadas.append(tarea)
            
            print(f"✓ Procesadas {len(tareas_procesadas)} tareas con descuentos aplicados")
            
            return tareas_procesadas
            
        except Error as e:
            print(f"✗ Error al extraer tareas: {e}")
            return []
    
    def clasificar_tipo_origen(self, tipo_actividad: str) -> str:
        """
        Clasifica el tipo de actividad en categorías generales
        
        Args:
            tipo_actividad: String con el tipo de actividad
            
        Returns:
            Categoría clasificada
        """
        if not tipo_actividad:
            return 'OTRA'
        
        tipo_lower = str(tipo_actividad).lower()
        
        # Cierres
        if 'cierre' in tipo_lower:
            return 'CIERRE'
        
        # Incidentes
        if 'incidente' in tipo_lower or 'soporte' in tipo_lower or 'correctivo' in tipo_lower:
            return 'INCIDENTE'
        
        # Implementaciones / Instalaciones / Aperturas
        if any(x in tipo_lower for x in ['implementacion', 'instalacion', 'apertura', 'migracion']):
            return 'IMPLEMENTACION'
        
        # POS / Datafonos / Envíos
        if any(x in tipo_lower for x in ['pos', 'datafono', 'envio', 'retiro']):
            return 'POS'
        
        # Órdenes de cambio
        if 'oc' in tipo_lower or 'orden de cambio' in tipo_lower or 'orden' in tipo_lower:
            return 'ORDEN_CAMBIO'
        
        # Visitas / Rollos
        if 'rollo' in tipo_lower or 'visita' in tipo_lower:
            return 'VISITA'
        
        return 'OTRA'
    
    def procesar_tareas(self, tareas: List[Dict]) -> Dict:
        """
        Procesa lista de tareas y las agrupa por técnico
        
        Args:
            tareas: Lista de diccionarios con tareas
            
        Returns:
            Diccionario con datos agrupados por técnico
        """
        tecnicos_data = {}
        tareas_sin_fecha = 0
        tareas_procesadas = 0
        
        for tarea in tareas:
            tecnico = str(tarea.get('tecnico', '')).strip()
            if not tecnico or tecnico.upper() == 'NONE':
                continue
            
            tipologia = str(tarea.get('tipologia', '') or '').strip().upper()
            
            # Obtener fecha de cierre
            fecha_cierre = tarea.get('fecha_cierre_plataforma_cliente')
            if not fecha_cierre:
                fecha_cierre = tarea.get('fecha_fin')
            
            if not fecha_cierre:
                tareas_sin_fecha += 1
                continue
            
            # Si es string, intentar convertir
            if isinstance(fecha_cierre, str):
                try:
                    fecha_cierre = datetime.strptime(fecha_cierre, '%Y-%m-%d %H:%M:%S')
                except:
                    tareas_sin_fecha += 1
                    continue
            
            mes = fecha_cierre.month
            anio = fecha_cierre.year
            mes_nombre = fecha_cierre.strftime('%B %Y')
            
            # Clasificar tipo de origen
            tipo_actividad = tarea.get('tipo_actividad', '')
            tipo_origen = self.clasificar_tipo_origen(tipo_actividad)
            
            # Los valores YA VIENEN calculados con descuentos
            valor_neto = float(tarea.get('prod_tecnico_final', 0) or 0)
            valor_bruto = float(tarea.get('valor_bruto', 0) or valor_neto)
            descuento = float(tarea.get('descuento_aplicado', 0) or 0)
            
            if valor_neto <= 0:
                continue
            
            # Obtener tipología y porcentaje (para registro)
            porcentaje_descuento = DESCUENTOS.get(tipologia, 0)
            
            # Inicializar técnico si no existe
            if tecnico not in tecnicos_data:
                tecnicos_data[tecnico] = {
                    'nombre': tecnico,
                    'meses': {},
                    'total_general': 0,
                    'total_tareas': 0,
                    'por_tipo_origen': {}
                }
            
            # Inicializar mes si no existe
            if mes_nombre not in tecnicos_data[tecnico]['meses']:
                tecnicos_data[tecnico]['meses'][mes_nombre] = {
                    'mes': mes_nombre,
                    'mes_numero': mes,
                    'anio': anio,
                    'tareas': [],
                    'resumen_tipologias': {},
                    'por_tipo_origen': {},
                    'total_bruto': 0,
                    'total_descuentos': 0,
                    'total_neto': 0,
                    'cantidad_tareas': 0
                }
            
            mes_data = tecnicos_data[tecnico]['meses'][mes_nombre]
            
            # Crear registro de tarea
            tarea_registro = {
                'tarea': tarea.get('tarea', ''),
                'formulario': tarea.get('formulario', ''),
                'tipo_origen': tipo_origen,
                'tipo_actividad': tipo_actividad,
                'tipificacion': tipologia,
                'trayecto': tarea.get('trayecto', ''),
                'valor_bruto': round(valor_bruto, 2),
                'porcentaje_descuento': porcentaje_descuento * 100,
                'valor_descuento': round(descuento, 2),
                'valor_neto': round(valor_neto, 2),
                'fecha_cierre': fecha_cierre.strftime('%Y-%m-%d %H:%M:%S'),
                'ciudad': tarea.get('ciudad', ''),
                'departamento': tarea.get('departamento', ''),
                'bodega': tarea.get('bodega', 'SIN ZONA'),
                'region': tarea.get('region_sitio', ''),
                'nombre_punto': tarea.get('nombre_punto', ''),
                'estado_ta': tarea.get('estado_ta', ''),
                'estado_fo': tarea.get('estado_fo', ''),
                'resultado': tarea.get('resultado_actividad', '')
            }
            
            # Agregar tarea
            mes_data['tareas'].append(tarea_registro)
            mes_data['total_bruto'] += valor_bruto
            mes_data['total_descuentos'] += descuento
            mes_data['total_neto'] += valor_neto
            mes_data['cantidad_tareas'] += 1
            
            # Actualizar contadores por tipo de origen (mes)
            if tipo_origen not in mes_data['por_tipo_origen']:
                mes_data['por_tipo_origen'][tipo_origen] = {'cantidad': 0, 'total': 0}
            mes_data['por_tipo_origen'][tipo_origen]['cantidad'] += 1
            mes_data['por_tipo_origen'][tipo_origen]['total'] += valor_neto
            
            # Actualizar contadores por tipo de origen (técnico)
            if tipo_origen not in tecnicos_data[tecnico]['por_tipo_origen']:
                tecnicos_data[tecnico]['por_tipo_origen'][tipo_origen] = {'cantidad': 0, 'total': 0}
            tecnicos_data[tecnico]['por_tipo_origen'][tipo_origen]['cantidad'] += 1
            tecnicos_data[tecnico]['por_tipo_origen'][tipo_origen]['total'] += valor_neto
            
            # Actualizar resumen por tipología
            if tipologia and tipologia != '':
                if tipologia not in mes_data['resumen_tipologias']:
                    mes_data['resumen_tipologias'][tipologia] = {
                        'cantidad': 0,
                        'total_bruto': 0,
                        'total_neto': 0,
                        'porcentaje_descuento': porcentaje_descuento * 100
                    }
                
                mes_data['resumen_tipologias'][tipologia]['cantidad'] += 1
                mes_data['resumen_tipologias'][tipologia]['total_bruto'] += valor_bruto
                mes_data['resumen_tipologias'][tipologia]['total_neto'] += valor_neto
            
            # Actualizar totales generales del técnico
            tecnicos_data[tecnico]['total_general'] += valor_neto
            tecnicos_data[tecnico]['total_tareas'] += 1
            
            tareas_procesadas += 1
        
        print(f"  ✓ Procesadas {tareas_procesadas} tareas del año 2026")
        if tareas_sin_fecha > 0:
            print(f"  ⚠ {tareas_sin_fecha} tareas sin fecha de cierre (excluidas)")
        
        return tecnicos_data
    
    def generar_resumen_global(self, tecnicos_data: Dict) -> Dict:
        """
        Genera un resumen global con estadísticas
        
        Args:
            tecnicos_data: Diccionario con datos de técnicos
            
        Returns:
            Diccionario con resumen global
        """
        total_tecnicos = len(tecnicos_data)
        total_tareas = sum(t['total_tareas'] for t in tecnicos_data.values())
        total_pagado = sum(t['total_general'] for t in tecnicos_data.values())
        
        # Contar por tipo de origen
        tipo_origen_global = {}
        for tecnico_data in tecnicos_data.values():
            for tipo, datos in tecnico_data['por_tipo_origen'].items():
                if tipo not in tipo_origen_global:
                    tipo_origen_global[tipo] = {'cantidad': 0, 'total': 0}
                tipo_origen_global[tipo]['cantidad'] += datos['cantidad']
                tipo_origen_global[tipo]['total'] += datos['total']
        
        return {
            'total_tecnicos': total_tecnicos,
            'total_tareas': total_tareas,
            'total_pagado': round(total_pagado, 2),
            'por_tipo_origen': tipo_origen_global,
            'fecha_generacion': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }
    
    def exportar_json(self, tecnicos_data: Dict, ruta_salida: str):
        """
        Exporta los datos a un archivo JSON (y .json.gz comprimido)
        
        Args:
            tecnicos_data: Diccionario con datos procesados
            ruta_salida: Ruta del archivo de salida
        """
        try:
            # Generar resumen global
            resumen = self.generar_resumen_global(tecnicos_data)
            
            # Crear estructura para JSON
            # Convertir diccionario de técnicos a lista
            tecnicos_lista = []
            for nombre, data in tecnicos_data.items():
                # Convertir diccionario de meses a lista
                meses_lista = []
                for mes_nombre, mes_data in data['meses'].items():
                    meses_lista.append(mes_data)
                
                tecnicos_lista.append({
                    'nombre': data['nombre'],
                    'meses': meses_lista,
                    'total_general': round(data['total_general'], 2),
                    'total_tareas': data['total_tareas'],
                    'por_tipo_origen': data['por_tipo_origen']
                })
            
            output_data = {
                'resumen': resumen,
                'tecnicos': tecnicos_lista
            }
            
            # Guardar JSON sin comprimir
            with open(ruta_salida, 'w', encoding='utf-8') as f:
                json.dump(output_data, f, ensure_ascii=False, indent=2)
            
            print(f"✓ Archivo JSON generado: {ruta_salida}")
            
            # Guardar JSON comprimido con gzip
            ruta_gz = ruta_salida + '.gz'
            json_str = json.dumps(output_data, ensure_ascii=False)
            json_bytes = json_str.encode('utf-8')
            
            with gzip.open(ruta_gz, 'wb') as f:
                f.write(json_bytes)
            
            # Mostrar tamaños
            tamano_json = os.path.getsize(ruta_salida) / 1024 / 1024  # MB
            tamano_gz = os.path.getsize(ruta_gz) / 1024 / 1024  # MB
            
            print(f"✓ Archivo JSON comprimido: {ruta_gz}")
            print(f"  Tamaño JSON: {tamano_json:.2f} MB")
            print(f"  Tamaño GZ: {tamano_gz:.2f} MB ({(tamano_gz/tamano_json*100):.1f}% del original)")
            
        except Exception as e:
            print(f"✗ Error al exportar JSON: {e}")
            import traceback
            traceback.print_exc()
    
    def generar_excel_resumen_comisiones(self, tecnicos_data: Dict, todas_tareas: List[Dict], ruta_salida: str):
        """
        Genera un Excel con el resumen de comisiones para ENERO 2026
        
        CORRECCIÓN: Usa los valores del JSON (que ya tienen descuentos aplicados correctamente)
        
        Args:
            tecnicos_data: Diccionario con datos de técnicos
            todas_tareas: Lista de todas las tareas (para detalles)
            ruta_salida: Ruta del archivo Excel de salida
        """
        try:
            # Crear workbook
            wb = Workbook()
            
            # Estilos comunes
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
            total_font = Font(bold=True)
            total_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # HOJA 1: Resumen de Comisiones por Técnico
            ws_resumen = wb.active
            ws_resumen.title = "Resumen Comisiones Enero"
            
            # Encabezados
            headers = ['TÉCNICO', 'TOTAL TAREAS', 'TOTAL BRUTO', 'TOTAL DESCUENTOS', 
                      'TOTAL NETO', 'META MENSUAL', 'EXCEDENTE', 'TIER', 'COMISIÓN']
            ws_resumen.append(headers)
            
            # Procesar cada técnico
            total_general_bruto = 0
            total_general_descuentos = 0
            total_general_neto = 0
            total_general_comision = 0
            
            for nombre_tecnico, tecnico_data in sorted(tecnicos_data.items()):
                # Buscar mes de enero
                enero_data = None
                for mes_nombre, mes_data in tecnico_data['meses'].items():
                    if mes_data.get('mes_numero') == 1 and mes_data.get('anio') == 2026:
                        enero_data = mes_data
                        break
                
                if enero_data:
                    total_bruto = enero_data.get('total_bruto', 0)
                    total_descuentos = enero_data.get('total_descuentos', 0)
                    total_neto = enero_data.get('total_neto', 0)
                    cantidad_tareas = enero_data.get('cantidad_tareas', 0)
                    
                    # Calcular comisión usando el sistema de tramos progresivos
                    comision_data = self.calcular_comision(total_neto)
                    comision = comision_data['commission']
                    excedente = comision_data['excedente']
                    tier_name = comision_data['tier']['name']
                    meta_cumplida = '✓' if comision_data['meta_cumplida'] else '✗'
                    
                    ws_resumen.append([
                        tecnico_data['nombre'],
                        cantidad_tareas,
                        total_bruto,
                        total_descuentos,
                        total_neto,
                        META_MENSUAL,
                        excedente,
                        f"{meta_cumplida} {tier_name}",
                        comision
                    ])
                    
                    total_general_bruto += total_bruto
                    total_general_descuentos += total_descuentos
                    total_general_neto += total_neto
                    total_general_comision += comision
            
            # Fila de totales
            ws_resumen.append([
                'TOTAL GENERAL',
                '',
                total_general_bruto,
                total_general_descuentos,
                total_general_neto,
                '',
                '',
                '',
                total_general_comision
            ])
            
            # Formatear hoja resumen
            for cell in ws_resumen[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            # Formatear última fila (totales)
            for cell in ws_resumen[ws_resumen.max_row]:
                cell.fill = total_fill
                cell.font = total_font
                cell.border = border
            
            # Formatear números
            for row in ws_resumen.iter_rows(min_row=2, max_row=ws_resumen.max_row):
                for idx, cell in enumerate(row):
                    if idx >= 2:  # Columnas numéricas
                        cell.number_format = '#,##0.00'
                    cell.border = border
            
            # Ajustar anchos
            ws_resumen.column_dimensions['A'].width = 40
            ws_resumen.column_dimensions['B'].width = 15
            ws_resumen.column_dimensions['C'].width = 18
            ws_resumen.column_dimensions['D'].width = 20
            ws_resumen.column_dimensions['E'].width = 18
            ws_resumen.column_dimensions['F'].width = 18
            ws_resumen.column_dimensions['G'].width = 18
            ws_resumen.column_dimensions['H'].width = 15
            ws_resumen.column_dimensions['I'].width = 18
            
            # HOJA 2: Detalle de Tareas con Descuentos
            ws_tareas = wb.create_sheet("Detalle Tareas Enero")
            
            headers_tareas = ['TÉCNICO', 'TAREA', 'TIPO ACTIVIDAD', 'TIPOLOGÍA', 
                             'VALOR BRUTO', 'DESCUENTO %', 'DESCUENTO $', 
                             'VALOR NETO', 'CIUDAD', 'FECHA']
            ws_tareas.append(headers_tareas)
            
            # Filtrar solo tareas de enero
            tareas_enero = [t for t in todas_tareas 
                           if t.get('fecha_cierre_plataforma_cliente') and 
                           t['fecha_cierre_plataforma_cliente'].month == 1]
            
            # Ordenar por técnico
            tareas_enero_ordenadas = sorted(tareas_enero, 
                                           key=lambda x: str(x.get('tecnico') or ''))
            
            for tarea in tareas_enero_ordenadas:
                tipologia = tarea.get('tipologia', '')
                descuento_porcentaje = DESCUENTOS.get(tipologia, 0) * 100
                
                ws_tareas.append([
                    tarea.get('tecnico', ''),
                    tarea.get('tarea', ''),
                    tarea.get('tipo_actividad', ''),
                    tipologia,
                    tarea.get('valor_bruto', 0),
                    descuento_porcentaje,
                    tarea.get('descuento_aplicado', 0),
                    tarea.get('prod_tecnico_final', 0),
                    tarea.get('ciudad', ''),
                    tarea.get('fecha_cierre_plataforma_cliente', '').strftime('%Y-%m-%d') 
                        if tarea.get('fecha_cierre_plataforma_cliente') else ''
                ])
            
            # Formatear hoja tareas
            for cell in ws_tareas[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            # Formatear números
            for row in ws_tareas.iter_rows(min_row=2, max_row=ws_tareas.max_row):
                for idx, cell in enumerate(row):
                    if idx in [4, 5, 6, 7]:
                        if idx == 5:  # Porcentaje
                            cell.number_format = '0.00"%"'
                        else:
                            cell.number_format = '#,##0.00'
                    cell.border = border
            
            # Ajustar anchos
            ws_tareas.column_dimensions['A'].width = 30
            ws_tareas.column_dimensions['B'].width = 15
            ws_tareas.column_dimensions['C'].width = 30
            ws_tareas.column_dimensions['D'].width = 20
            for col in ['E', 'F', 'G', 'H']:
                ws_tareas.column_dimensions[col].width = 18
            ws_tareas.column_dimensions['I'].width = 20
            ws_tareas.column_dimensions['J'].width = 15
            
            # HOJA 3: Resumen por Tipología
            ws_tipologia = wb.create_sheet("Resumen por Tipología")
            
            headers_tipologia = ['TIPOLOGÍA', 'DESCUENTO %', 'CANTIDAD TAREAS', 
                                'TOTAL BRUTO', 'TOTAL DESCUENTOS', 'TOTAL NETO']
            ws_tipologia.append(headers_tipologia)
            
            # Calcular totales por tipología
            tipologia_stats = {}
            for tarea in tareas_enero:
                tipo = tarea.get('tipologia', 'SIN TIPOLOGÍA')
                if tipo not in tipologia_stats:
                    tipologia_stats[tipo] = {
                        'cantidad': 0,
                        'bruto': 0,
                        'descuentos': 0,
                        'neto': 0
                    }
                
                tipologia_stats[tipo]['cantidad'] += 1
                tipologia_stats[tipo]['bruto'] += tarea.get('valor_bruto', 0)
                tipologia_stats[tipo]['descuentos'] += tarea.get('descuento_aplicado', 0)
                tipologia_stats[tipo]['neto'] += tarea.get('prod_tecnico_final', 0)
            
            for tipo, stats in sorted(tipologia_stats.items()):
                descuento_pct = DESCUENTOS.get(tipo, 0) * 100
                ws_tipologia.append([
                    tipo,
                    descuento_pct,
                    stats['cantidad'],
                    stats['bruto'],
                    stats['descuentos'],
                    stats['neto']
                ])
            
            # Formatear
            for cell in ws_tipologia[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            for row in ws_tipologia.iter_rows(min_row=2, max_row=ws_tipologia.max_row):
                for idx, cell in enumerate(row):
                    if idx == 1:  # Porcentaje
                        cell.number_format = '0.00"%"'
                    elif idx >= 2:
                        cell.number_format = '#,##0.00'
                    cell.border = border
            
            ws_tipologia.column_dimensions['A'].width = 25
            ws_tipologia.column_dimensions['B'].width = 15
            for col in ['C', 'D', 'E', 'F']:
                ws_tipologia.column_dimensions[col].width = 18
            
            # HOJA 4: Explicación del Sistema de Comisiones
            ws_sistema = wb.create_sheet("Sistema de Comisiones")
            
            # Título
            ws_sistema.merge_cells('A1:D1')
            ws_sistema['A1'] = 'SISTEMA DE COMISIONES POR TRAMOS PROGRESIVOS'
            ws_sistema['A1'].font = Font(bold=True, size=14, color="FFFFFF")
            ws_sistema['A1'].fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
            ws_sistema['A1'].alignment = Alignment(horizontal='center')
            
            ws_sistema.append([''])  # Fila vacía
            
            # Descripción
            ws_sistema.merge_cells('A3:D3')
            ws_sistema['A3'] = 'Cómo funciona:'
            ws_sistema['A3'].font = Font(bold=True, size=12)
            
            ws_sistema.merge_cells('A4:D5')
            ws_sistema['A4'] = ('1. Se calcula el EXCEDENTE: Total Neto - Meta Mensual ($4,500,000)\n'
                               '2. Se aplican los porcentajes de comisión a cada TRAMO del excedente\n'
                               '3. Se SUMAN todas las comisiones de cada tramo')
            ws_sistema['A4'].alignment = Alignment(wrap_text=True, vertical='top')
            
            ws_sistema.append([''])  # Fila vacía
            
            # Tabla de tramos
            ws_sistema.append(['TRAMO', 'RANGO', 'COMISIÓN'])
            for cell in ws_sistema[ws_sistema.max_row]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            
            for tier in COMMISSION_TIERS:
                if tier['max'] == float('inf'):
                    rango = f"Sobre ${tier['min']:,}"
                else:
                    rango = f"${tier['min']:,} - ${tier['max']:,}"
                ws_sistema.append([
                    f"Tramo {tier['name']}",
                    rango,
                    f"{tier['percentage']*100}%"
                ])
            
            ws_sistema.append([''])  # Fila vacía
            
            # EJEMPLO 1: Excedente de $1,500,000
            ws_sistema.merge_cells(f'A{ws_sistema.max_row + 1}:D{ws_sistema.max_row + 1}')
            ws_sistema[f'A{ws_sistema.max_row}'] = 'EJEMPLO 1: Técnico con excedente de $1,500,000'
            ws_sistema[f'A{ws_sistema.max_row}'].font = Font(bold=True, size=11)
            ws_sistema[f'A{ws_sistema.max_row}'].fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
            
            ejemplo1_excedente = 1500000
            ejemplo_data1 = self.calcular_comision(META_MENSUAL + ejemplo1_excedente)
            
            ws_sistema.append(['Concepto', 'Cálculo', '', 'Resultado'])
            ws_sistema.append(['Total Neto:', f"${META_MENSUAL + ejemplo1_excedente:,.0f}", '', ''])
            ws_sistema.append(['Meta Mensual:', f"${META_MENSUAL:,.0f}", '', ''])
            ws_sistema.append(['Excedente:', f"${ejemplo1_excedente:,.0f}", '', ''])
            ws_sistema.append(['', '', '', ''])
            ws_sistema.append(['CÁLCULO DE COMISIÓN:', '', '', ''])
            
            # Tramo 1
            tramo1_1 = 1000000
            com1_1 = tramo1_1 * 0.15
            ws_sistema.append(['  Tramo 1 (15%):', f"${tramo1_1:,.0f} × 15%", '=', f"${com1_1:,.0f}"])
            
            # Tramo 2
            tramo2_1 = 500000
            com2_1 = tramo2_1 * 0.08
            ws_sistema.append(['  Tramo 2 (8%):', f"${tramo2_1:,.0f} × 8%", '=', f"${com2_1:,.0f}"])
            
            ws_sistema.append(['', '', '', ''])
            ws_sistema.append(['COMISIÓN TOTAL:', '', '=', f"${ejemplo_data1['commission']:,.0f}"])
            
            # Formatear fila de total
            ultima_fila = ws_sistema.max_row
            for cell in ws_sistema[ultima_fila]:
                cell.font = Font(bold=True, size=11)
                cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
            
            ws_sistema.append([''])  # Fila vacía
            
            # EJEMPLO 2: Excedente de $4,200,000
            ws_sistema.merge_cells(f'A{ws_sistema.max_row + 1}:D{ws_sistema.max_row + 1}')
            ws_sistema[f'A{ws_sistema.max_row}'] = 'EJEMPLO 2: Técnico con excedente de $4,200,000'
            ws_sistema[f'A{ws_sistema.max_row}'].font = Font(bold=True, size=11)
            ws_sistema[f'A{ws_sistema.max_row}'].fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
            
            ejemplo2_excedente = 4200000
            ejemplo_data2 = self.calcular_comision(META_MENSUAL + ejemplo2_excedente)
            
            ws_sistema.append(['Concepto', 'Cálculo', '', 'Resultado'])
            ws_sistema.append(['Total Neto:', f"${META_MENSUAL + ejemplo2_excedente:,.0f}", '', ''])
            ws_sistema.append(['Meta Mensual:', f"${META_MENSUAL:,.0f}", '', ''])
            ws_sistema.append(['Excedente:', f"${ejemplo2_excedente:,.0f}", '', ''])
            ws_sistema.append(['', '', '', ''])
            ws_sistema.append(['CÁLCULO DE COMISIÓN:', '', '', ''])
            
            # Tramo 1
            tramo1_2 = 1000000
            com1_2 = tramo1_2 * 0.15
            ws_sistema.append(['  Tramo 1 (15%):', f"${tramo1_2:,.0f} × 15%", '=', f"${com1_2:,.0f}"])
            
            # Tramo 2
            tramo2_2 = 1000000
            com2_2 = tramo2_2 * 0.08
            ws_sistema.append(['  Tramo 2 (8%):', f"${tramo2_2:,.0f} × 8%", '=', f"${com2_2:,.0f}"])
            
            # Tramo 3
            tramo3_2 = 1000000
            com3_2 = tramo3_2 * 0.05
            ws_sistema.append(['  Tramo 3 (5%):', f"${tramo3_2:,.0f} × 5%", '=', f"${com3_2:,.0f}"])
            
            # Tramo 4
            tramo4_2 = ejemplo2_excedente - 3000000
            com4_2 = tramo4_2 * 0.03
            ws_sistema.append(['  Tramo 4 (3%):', f"${tramo4_2:,.0f} × 3%", '=', f"${com4_2:,.0f}"])
            
            ws_sistema.append(['', '', '', ''])
            ws_sistema.append(['COMISIÓN TOTAL:', '', '=', f"${ejemplo_data2['commission']:,.0f}"])
            
            # Formatear fila de total
            ultima_fila2 = ws_sistema.max_row
            for cell in ws_sistema[ultima_fila2]:
                cell.font = Font(bold=True, size=11)
                cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
            
            # Nota importante
            ws_sistema.append([''])  # Fila vacía
            ws_sistema.merge_cells(f'A{ws_sistema.max_row + 1}:D{ws_sistema.max_row + 1}')
            ws_sistema[f'A{ws_sistema.max_row}'] = '⚠️ IMPORTANTE: Se suman las comisiones de cada tramo alcanzado'
            ws_sistema[f'A{ws_sistema.max_row}'].font = Font(italic=True, size=10, color="0000FF")
            ws_sistema[f'A{ws_sistema.max_row}'].alignment = Alignment(horizontal='center')
            
            # Ajustar anchos de columnas
            ws_sistema.column_dimensions['A'].width = 25
            ws_sistema.column_dimensions['B'].width = 25
            ws_sistema.column_dimensions['C'].width = 10
            ws_sistema.column_dimensions['D'].width = 20
            
            # Guardar archivo
            wb.save(ruta_salida)
            print(f"✓ Excel generado exitosamente: {ruta_salida}")
            print(f"  - Hoja 1: Resumen de comisiones por técnico (tramos progresivos)")
            print(f"  - Hoja 2: Detalle de {len(tareas_enero)} tareas de enero")
            print(f"  - Hoja 3: Resumen por tipología")
            print(f"  - Hoja 4: Explicación del sistema de comisiones (con ejemplos)")
            print(f"\n  Meta mensual: ${META_MENSUAL:,.0f}")
            print(f"  Tramos de comisión (progresivos - se suman):")
            for tier in COMMISSION_TIERS:
                print(f"    Tramo {tier['name']}: {tier['percentage']*100}% (${tier['min']:,.0f} - {'∞' if tier['max'] == float('inf') else f'${tier['max']:,.0f}'})")
            
        except Exception as e:
            print(f"✗ Error al generar Excel: {e}")
            import traceback
            traceback.print_exc()
    
    
    def mostrar_resumen_tecnico(self, tecnicos_data: Dict, nombre_tecnico: str):
        """
        Muestra un resumen de un técnico específico
        
        Args:
            tecnicos_data: Diccionario con datos de técnicos
            nombre_tecnico: Nombre del técnico a consultar
        """
        # Buscar técnico (case insensitive)
        tecnico_encontrado = None
        for nombre, data in tecnicos_data.items():
            if nombre_tecnico.upper() in nombre.upper():
                tecnico_encontrado = data
                break
        
        if not tecnico_encontrado:
            print(f"\n✗ Técnico '{nombre_tecnico}' no encontrado")
            print("\nTécnicos disponibles (primeros 10):")
            for i, nombre in enumerate(list(tecnicos_data.keys())[:10]):
                print(f"  {i+1}. {nombre}")
            return
        
        print(f"\n{'='*80}")
        print(f"RESUMEN PARA: {tecnico_encontrado['nombre']} (AÑO 2026)")
        print(f"{'='*80}")
        print(f"Total general: ${tecnico_encontrado['total_general']:,.2f}")
        print(f"Total tareas: {tecnico_encontrado['total_tareas']}")
        
        print(f"\nPor tipo de origen:")
        for tipo, datos in sorted(tecnico_encontrado['por_tipo_origen'].items()):
            if datos['cantidad'] > 0:
                print(f"  {tipo}: {datos['cantidad']} tareas - ${datos['total']:,.2f}")
        
        print(f"\nDetalle por mes (últimos 3 meses):")
        print(f"{'-'*80}")
        
        # Mostrar solo los últimos 3 meses
        meses_ordenados = sorted(tecnico_encontrado['meses'].items(), 
                                key=lambda x: (x[1]['anio'], x[1]['mes_numero']), 
                                reverse=True)
        
        for mes_nombre, mes_data in meses_ordenados[:3]:
            print(f"\n{mes_nombre}:")
            print(f"  Tareas: {mes_data['cantidad_tareas']}")
            print(f"  Total bruto: ${mes_data['total_bruto']:,.2f}")
            print(f"  Descuentos: ${mes_data['total_descuentos']:,.2f}")
            print(f"  Total neto: ${mes_data['total_neto']:,.2f}")
            
            if mes_data['por_tipo_origen']:
                print(f"\n  Por tipo de origen:")
                for tipo, datos in sorted(mes_data['por_tipo_origen'].items()):
                    if datos['cantidad'] > 0:
                        print(f"    {tipo}: {datos['cantidad']} tareas - ${datos['total']:,.2f}")
            
            if mes_data['resumen_tipologias']:
                print(f"\n  Por tipología:")
                for tipo, datos in sorted(mes_data['resumen_tipologias'].items()):
                    print(f"    {tipo}: {datos['cantidad']} tareas - ${datos['total_neto']:,.2f} (descuento {datos['porcentaje_descuento']}%)")
        
        print(f"\n{'='*80}")


def main():
    """Función principal"""
    print("="*80)
    print("EXTRACTOR DE LIQUIDACIONES - LINEACOM")
    print("ENERO 2026: Desde Excel | FEBRERO+ 2026: Desde Base de Datos")
    print("="*80)
    
    # Crear extractor
    extractor = ExtractorLiquidacionesDB(DB_CONFIG)
    
    # Conectar a la base de datos
    print("\n1. Conectando a la base de datos...")
    if not extractor.conectar():
        print("✗ No se pudo conectar a la base de datos. Abortando.")
        return
    
    try:
        # Mostrar tablas disponibles
        print("\n2. Verificando tablas necesarias...")
        tablas = extractor.obtener_tablas_disponibles()
        tablas_necesarias = ['data_linea_todos', 'facturacion_linea']
        for tabla in tablas_necesarias:
            if tabla in tablas:
                print(f"  ✓ Tabla {tabla} encontrada")
            else:
                print(f"  ✗ Tabla {tabla} no encontrada")
                return
        
        # Extraer datos de ENERO desde Excel
        print("\n3. Extrayendo tareas de ENERO 2026 desde Excel...")
        tareas_enero = extractor.extraer_tareas_enero_desde_excel()
        
        
        # Extraer datos de FEBRERO+ desde base de datos
        print("\n4. Extrayendo tareas de FEBRERO-DICIEMBRE 2026 desde base de datos...")
        tareas_db = extractor.extraer_tareas()
        
        # Combinar todas las tareas
        todas_tareas = tareas_enero + tareas_db
        total_tareas = len(todas_tareas)
        
        if not todas_tareas:
            print("✗ No se encontraron tareas del año 2026")
            return
        
        print(f"\n✓ Total de tareas combinadas: {total_tareas}")
        print(f"  - Enero (Excel): {len(tareas_enero)}")
        print(f"  - Feb-Dic (DB): {len(tareas_db)}")
        
        # Procesar tareas
        print("\n5. Procesando tareas y agrupando por técnico...")
        tecnicos_data = extractor.procesar_tareas(todas_tareas)
        print(f"  ✓ Procesados datos de {len(tecnicos_data)} técnicos")
        
        # Exportar a JSON
        print("\n6. Generando archivos JSON...")
        ruta_json = os.path.join(os.getcwd(), 'liquidaciones_db.json')
        extractor.exportar_json(tecnicos_data, ruta_json)
        
        # Generar Excel de resumen de comisiones
        print("\n7. Generando Excel de resumen de comisiones de enero...")
        ruta_excel = os.path.join(os.getcwd(), 'resumen_comisiones_enero_2026.xlsx')
        extractor.generar_excel_resumen_comisiones(tecnicos_data, todas_tareas, ruta_excel)
        
        # Mostrar ejemplo con un técnico
        if tecnicos_data:
            print("\n8. Ejemplo de consulta:")
            primer_tecnico = list(tecnicos_data.keys())[0]
            extractor.mostrar_resumen_tecnico(tecnicos_data, primer_tecnico.split()[0])
        
    finally:
        # Desconectar de la base de datos
        print("\n9. Cerrando conexión...")
        extractor.desconectar()
    
    print("\n" + "="*80)
    print("PROCESO COMPLETADO - DATOS DE 2026")
    print("Archivos generados:")
    print("  - liquidaciones_db.json / .json.gz")
    print("  - resumen_comisiones_enero_2026.xlsx")
    print("="*80)

if __name__ == "__main__":
    main()